import io
import csv
import logging
from services.storage_service import StorageService

logger = logging.getLogger(__name__)

# Maximum bytes to download for extraction (20 MB)
MAX_EXTRACT_SIZE = 20 * 1024 * 1024
# Maximum extracted text length (50k chars)
MAX_TEXT_LENGTH = 50_000


class ContentExtraction:
    """Extract text from uploaded files for RAG indexing"""

    @classmethod
    def extract_from_storage(cls, file_path, file_type):
        """Download a file from Firebase Storage and extract text.

        Args:
            file_path: Path in Firebase Storage (e.g. 'content/abc123.pdf')
            file_type: Content type string ('PDF', 'CSV', 'Excel', 'Text', 'Document')

        Returns:
            str: Extracted text, or empty string on failure
        """
        try:
            bucket = StorageService.get_bucket()
            blob = bucket.blob(file_path)
            # Check size before downloading to prevent memory exhaustion
            blob.reload()
            if blob.size and blob.size > MAX_EXTRACT_SIZE:
                logger.warning(f"File too large for extraction: {file_path} ({blob.size} bytes)")
                return ''
            file_bytes = blob.download_as_bytes()

            file_type_upper = (file_type or '').upper()

            if file_type_upper == 'PDF':
                return cls._extract_pdf(file_bytes)
            elif file_type_upper == 'CSV':
                return cls._extract_csv(file_bytes)
            elif file_type_upper == 'EXCEL':
                return cls._extract_excel(file_bytes)
            else:
                # Try to read as plain text
                return cls._extract_text(file_bytes)
        except Exception as e:
            logger.warning(f"Content extraction error for {file_path}: {e}")
            return ''

    @classmethod
    def _extract_pdf(cls, file_bytes):
        """Extract text from PDF bytes"""
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(io.BytesIO(file_bytes))
            pages = []
            total_len = 0
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    pages.append(text)
                    total_len += len(text)
                    if total_len > MAX_TEXT_LENGTH:
                        pages.append('... (truncated)')
                        break
            return '\n\n'.join(pages)[:MAX_TEXT_LENGTH]
        except Exception as e:
            logger.warning(f"PDF extraction failed: {e}")
            return ''

    @classmethod
    def _extract_csv(cls, file_bytes):
        """Extract text from CSV bytes"""
        try:
            text = file_bytes.decode('utf-8', errors='replace')
            reader = csv.reader(io.StringIO(text))
            rows = []
            for i, row in enumerate(reader):
                rows.append(', '.join(row))
                if i > 500:  # Limit to 500 rows
                    rows.append('... (truncated)')
                    break
            return '\n'.join(rows)
        except Exception as e:
            logger.warning(f"CSV extraction failed: {e}")
            return ''

    @classmethod
    def _extract_excel(cls, file_bytes):
        """Extract text from Excel bytes"""
        wb = None
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    cells = [str(c) if c is not None else '' for c in row[:100]]  # Cap columns
                    rows.append(', '.join(cells))
                    if i > 500:
                        rows.append('... (truncated)')
                        break
                sheets.append(f"[Sheet: {sheet_name}]\n" + '\n'.join(rows))
            return '\n\n'.join(sheets)[:MAX_TEXT_LENGTH]
        except Exception as e:
            logger.warning(f"Excel extraction failed: {e}")
            return ''
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception as close_err:
                    logger.debug(f"Error closing workbook: {close_err}")

    @classmethod
    def _extract_text(cls, file_bytes):
        """Try to read bytes as plain text"""
        try:
            return file_bytes.decode('utf-8', errors='replace')[:50000]
        except Exception as e:
            logger.warning(f"Text extraction failed: {e}")
            return ''
